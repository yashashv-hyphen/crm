import uuid
import os
import json
import logging
from datetime import date, datetime, timezone

import redis as sync_redis
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

from app.celery_app import celery_app
from app.config import settings
from app.utils.excel_validator import read_excel_rows
from app.services.s3_service import download_file_from_s3
from app.services.upload_service import get_dev_temp_path

logger = logging.getLogger(__name__)


def _publish_upload_complete(upload_file_id: str, success_count: int, error_count: int, fos_ids: list[str]) -> None:
    """Publish upload completion event to Redis pub/sub channel. Non-fatal if Redis is unreachable."""
    try:
        r = sync_redis.Redis.from_url(settings.redis_url)
        r.publish("crm:upload_complete", json.dumps({
            "type": "upload_complete",
            "upload_id": upload_file_id,
            "success_count": success_count,
            "error_count": error_count,
            "fos_ids": fos_ids,
        }))
        r.close()
    except Exception as exc:
        logger.warning("Redis publish failed (non-fatal): %s", exc)

# Sync engine for Celery (asyncpg → psycopg2)
_sync_db_url = settings.database_url.replace("postgresql+asyncpg", "postgresql+psycopg2")
_sync_engine = create_engine(_sync_db_url, pool_pre_ping=True)
SyncSession = sessionmaker(bind=_sync_engine)


def _get_file_bytes(upload_id: str, s3_key: str) -> bytes:
    if not settings.r2_bucket:
        path = get_dev_temp_path(upload_id)
        with open(path, "rb") as f:
            return f.read()
    return download_file_from_s3(s3_key)


def _update_upload_status(session: Session, upload_id: str, status: str, total: int, success: int, errors: int) -> None:
    session.execute(
        text("""
            UPDATE upload_files
            SET status = :status, total_rows = :total, success_rows = :success,
                error_rows = :errors, completed_at = :now
            WHERE id = :id
        """),
        {"status": status, "total": total, "success": success, "errors": errors,
         "now": datetime.now(timezone.utc), "id": upload_id},
    )


def _insert_error(session: Session, upload_file_id: str, row_number: int, merchant_id: str | None, error_type: str, detail: str) -> None:
    session.execute(
        text("""
            INSERT INTO upload_errors (id, upload_file_id, row_number, merchant_id, error_type, error_detail)
            VALUES (:id, :upload_file_id, :row_number, :merchant_id, :error_type, :error_detail)
        """),
        {
            "id": str(uuid.uuid4()),
            "upload_file_id": upload_file_id,
            "row_number": row_number,
            "merchant_id": merchant_id,
            "error_type": error_type,
            "error_detail": detail,
        },
    )


@celery_app.task(bind=True, name="process_template1_upload")
def process_template1_upload(self, upload_file_id: str) -> dict:
    with SyncSession() as session:
        try:
            # Get upload record
            row = session.execute(
                text("SELECT s3_key, activity_id, admin_id FROM upload_files WHERE id = :id"),
                {"id": upload_file_id},
            ).fetchone()
            if not row:
                return {"error": "Upload record not found"}

            s3_key, activity_id, admin_id = row.s3_key, row.activity_id, row.admin_id

            # Mark as processing
            session.execute(
                text("UPDATE upload_files SET status = 'processing' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()

            file_bytes = _get_file_bytes(upload_file_id, s3_key)
            _, rows = read_excel_rows(file_bytes)

            # Load existing MIDs from DB
            existing_mids_result = session.execute(text("SELECT merchant_id FROM leads"))
            existing_mids = {r[0] for r in existing_mids_result}

            seen_mids_in_file: set[str] = set()
            error_count = 0
            success_count = 0
            total = len(rows)

            # Get FOS user map: email → id
            fos_users = session.execute(
                text("SELECT id, email FROM users WHERE role = 'fos' AND is_active = true")
            ).fetchall()
            fos_map = {u.email.lower(): str(u.id) for u in fos_users}

            # Get activity id by name
            if activity_id:
                act_result = session.execute(
                    text("SELECT name FROM activities WHERE id = :id"), {"id": str(activity_id)}
                ).fetchone()
                activity_name = act_result[0] if act_result else None
            else:
                activity_name = None

            leads_to_insert = []

            for idx, row_data in enumerate(rows, start=2):
                merchant_id = str(row_data.get("Merchant ID") or "").strip()
                mobile = str(row_data.get("Mobile Number") or "").strip()
                assigned_to = str(row_data.get("Assigned To FOS") or "").strip().lower()
                seller_name = str(row_data.get("Seller Name") or "").strip() or None
                email_id = str(row_data.get("Email ID") or "").strip() or None
                remark = str(row_data.get("Remark") or "").strip() or None
                stage_assigned = str(row_data.get("Stage Assigned") or "").strip() or None
                week_no_raw = row_data.get("Week No")
                year_raw = row_data.get("Year")
                date_of_assignment_raw = row_data.get("Date of Assignment")

                if not merchant_id:
                    _insert_error(session, upload_file_id, idx, None, "missing_field", "Merchant ID is blank")
                    error_count += 1
                    continue

                if merchant_id in seen_mids_in_file:
                    _insert_error(session, upload_file_id, idx, merchant_id, "duplicate_in_file", f"Merchant ID {merchant_id} appears more than once in this file")
                    error_count += 1
                    continue
                seen_mids_in_file.add(merchant_id)

                if merchant_id in existing_mids:
                    _insert_error(session, upload_file_id, idx, merchant_id, "duplicate_mid", f"Merchant ID {merchant_id} already exists in the system")
                    error_count += 1
                    continue

                if mobile and (not mobile.isdigit() or len(mobile) != 10):
                    _insert_error(session, upload_file_id, idx, merchant_id, "invalid_mobile", f"Mobile number must be exactly 10 digits, got: {mobile}")
                    error_count += 1
                    continue

                fos_id = fos_map.get(assigned_to)
                if not fos_id:
                    _insert_error(session, upload_file_id, idx, merchant_id, "fos_not_found", f"FOS '{assigned_to}' not found in system")
                    error_count += 1
                    continue

                # Parse date
                doa = None
                if date_of_assignment_raw:
                    try:
                        if isinstance(date_of_assignment_raw, (date, datetime)):
                            doa = date_of_assignment_raw if isinstance(date_of_assignment_raw, date) else date_of_assignment_raw.date()
                        else:
                            from datetime import datetime as dt
                            doa = dt.strptime(str(date_of_assignment_raw).strip(), "%d/%m/%Y").date()
                    except Exception:
                        doa = None

                leads_to_insert.append({
                    "id": str(uuid.uuid4()),
                    "merchant_id": merchant_id,
                    "seller_name": seller_name,
                    "mobile_number": mobile or None,
                    "email_id": email_id,
                    "stage_assigned": stage_assigned or activity_name,
                    "date_of_assignment": doa,
                    "week_no": int(week_no_raw) if week_no_raw else None,
                    "year": int(year_raw) if year_raw else None,
                    "current_activity_id": str(activity_id) if activity_id else None,
                    "current_stage": activity_name,
                    "assigned_fos_id": fos_id,
                    "remark": remark,
                    "source_upload_id": upload_file_id,
                    "rnc_entry_date": date.today() if activity_name and "rnc" in activity_name.lower() else None,
                    "idv_entry_date": date.today() if activity_name and "idv" in activity_name.lower() else None,
                    "rtl_entry_date": date.today() if activity_name and "rtl" in activity_name.lower() else None,
                    "fba_entry_date": date.today() if activity_name and "fba" in activity_name.lower() else None,
                    "sp_entry_date": date.today() if activity_name and "sp" in activity_name.lower() else None,
                    "open_spending_entry_date": date.today() if activity_name and "open spending" in activity_name.lower() else None,
                    "narf_entry_date": date.today() if activity_name and "narf" in activity_name.lower() else None,
                    "gsi_entry_date": date.today() if activity_name and "gsi" in activity_name.lower() else None,
                })

            # Bulk insert valid leads
            if leads_to_insert:
                session.execute(
                    text("""
                        INSERT INTO leads (
                            id, merchant_id, seller_name, mobile_number, email_id,
                            stage_assigned, date_of_assignment, week_no, year,
                            current_activity_id, current_stage, assigned_fos_id, remark,
                            source_upload_id,
                            rnc_entry_date, idv_entry_date, rtl_entry_date, fba_entry_date,
                            sp_entry_date, open_spending_entry_date, narf_entry_date, gsi_entry_date
                        ) VALUES (
                            :id, :merchant_id, :seller_name, :mobile_number, :email_id,
                            :stage_assigned, :date_of_assignment, :week_no, :year,
                            :current_activity_id, :current_stage, :assigned_fos_id, :remark,
                            :source_upload_id,
                            :rnc_entry_date, :idv_entry_date, :rtl_entry_date, :fba_entry_date,
                            :sp_entry_date, :open_spending_entry_date, :narf_entry_date, :gsi_entry_date
                        )
                    """),
                    leads_to_insert,
                )
                success_count = len(leads_to_insert)

            _update_upload_status(session, upload_file_id, "completed", total, success_count, error_count)
            session.commit()

            if leads_to_insert:
                unique_fos_ids = list({lead["assigned_fos_id"] for lead in leads_to_insert})
                _publish_upload_complete(upload_file_id, success_count, error_count, unique_fos_ids)

            return {"total": total, "success": success_count, "errors": error_count}

        except Exception as exc:
            session.rollback()
            session.execute(
                text("UPDATE upload_files SET status = 'failed' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()
            raise exc


@celery_app.task(bind=True, name="process_regular_upload")
def process_regular_upload(self, upload_file_id: str) -> dict:
    """Parse the NTCPL Sheet3 format: merchant_customer_id, Lead Stage, GSE, Number, Email ID, Assigned on, Type."""
    with SyncSession() as session:
        try:
            row = session.execute(
                text("SELECT s3_key, admin_id FROM upload_files WHERE id = :id"),
                {"id": upload_file_id},
            ).fetchone()
            if not row:
                return {"error": "Upload record not found"}

            s3_key, admin_id = row.s3_key, row.admin_id

            session.execute(
                text("UPDATE upload_files SET status = 'processing' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()

            file_bytes = _get_file_bytes(upload_file_id, s3_key)

            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            # Find sheet with 'merchant_customer_id' header
            ws = None
            for sheet_name in wb.sheetnames:
                s = wb[sheet_name]
                headers = [str(c.value).strip() if c.value else "" for c in next(s.iter_rows(max_row=1))]
                if "merchant_customer_id" in headers:
                    ws = s
                    break
            if ws is None:
                session.execute(text("UPDATE upload_files SET status = 'failed' WHERE id = :id"), {"id": upload_file_id})
                session.commit()
                return {"error": "No sheet with merchant_customer_id header found"}

            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            col = {h: i for i, h in enumerate(headers)}

            def _get(row_vals, name):
                idx = col.get(name)
                if idx is None:
                    return None
                val = row_vals[idx]
                return val

            # Columns consumed into named Lead fields below — everything else in the
            # sheet (e.g. marketplace_id, is_reg_started, is_idv_passed, Priority) gets
            # preserved verbatim in custom_data instead of being silently dropped.
            _consumed_headers = {"merchant_customer_id", "Number", "Lead Stage", "GSE", "Email ID", "Type", "Assigned on"}
            _extra_headers = [h for h in headers if h and h not in _consumed_headers]

            def _json_safe(val):
                if val is None:
                    return None
                if isinstance(val, (date, datetime)):
                    return val.isoformat()
                if isinstance(val, (str, int, float, bool)):
                    return val
                return str(val)

            # Load all activities for Lead Stage → activity mapping
            activities_rows = session.execute(
                text("SELECT id, name FROM activities WHERE is_active = true")
            ).fetchall()
            activity_map = {a.name.lower(): str(a.id) for a in activities_rows}

            def _resolve_activity(stage: str):
                stage_lower = stage.lower()
                # exact match first
                if stage_lower in activity_map:
                    return next((str(a.id) for a in activities_rows if a.name.lower() == stage_lower), None), \
                           next((a.name for a in activities_rows if a.name.lower() == stage_lower), None)
                # substring: "rnc" matches "rnc pending"
                for a in activities_rows:
                    if stage_lower in a.name.lower():
                        return str(a.id), a.name
                return None, None

            # Load FOS users for GSE name matching
            fos_rows = session.execute(
                text("SELECT id, full_name FROM users WHERE role = 'fos' AND is_active = true")
            ).fetchall()

            def _resolve_fos(gse: str):
                gse_lower = gse.lower().strip()
                for u in fos_rows:
                    if u.full_name.lower().startswith(gse_lower):
                        return str(u.id)
                return None

            # Load existing MIDs
            existing_mids = {r[0] for r in session.execute(text("SELECT merchant_id FROM leads")).fetchall()}
            seen_mids: set[str] = set()

            error_count = 0
            success_count = 0
            leads_to_insert = []

            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len(all_rows)

            for idx, row_vals in enumerate(all_rows, start=2):
                merchant_id = str(_get(row_vals, "merchant_customer_id") or "").strip()
                if not merchant_id or merchant_id == "None":
                    _insert_error(session, upload_file_id, idx, None, "missing_field", "merchant_customer_id is blank")
                    error_count += 1
                    continue

                if merchant_id in seen_mids:
                    _insert_error(session, upload_file_id, idx, merchant_id, "duplicate_in_file", f"Merchant ID {merchant_id} appears more than once")
                    error_count += 1
                    continue
                seen_mids.add(merchant_id)

                if merchant_id in existing_mids:
                    _insert_error(session, upload_file_id, idx, merchant_id, "duplicate_mid", f"Merchant ID {merchant_id} already exists")
                    error_count += 1
                    continue

                # Mobile: strip country code prefix if 12+ digits
                raw_mobile = str(_get(row_vals, "Number") or "").strip()
                if raw_mobile and raw_mobile != "None":
                    digits = ''.join(c for c in raw_mobile if c.isdigit())
                    mobile = digits[-10:] if len(digits) >= 10 else digits
                    if mobile and (not mobile.isdigit() or len(mobile) != 10):
                        _insert_error(session, upload_file_id, idx, merchant_id, "invalid_mobile", f"Mobile must be 10 digits, got: {mobile}")
                        error_count += 1
                        continue
                else:
                    mobile = None

                lead_stage = str(_get(row_vals, "Lead Stage") or "").strip()
                if not lead_stage or lead_stage == "None":
                    _insert_error(session, upload_file_id, idx, merchant_id, "missing_field", "Lead Stage is blank")
                    error_count += 1
                    continue

                activity_id, activity_name = _resolve_activity(lead_stage)
                if not activity_id:
                    _insert_error(session, upload_file_id, idx, merchant_id, "activity_not_found", f"Lead Stage '{lead_stage}' does not match any activity")
                    error_count += 1
                    continue

                gse = str(_get(row_vals, "GSE") or "").strip()
                if not gse or gse == "None":
                    _insert_error(session, upload_file_id, idx, merchant_id, "missing_field", "GSE (FOS name) is blank")
                    error_count += 1
                    continue

                fos_id = _resolve_fos(gse)
                if not fos_id:
                    _insert_error(session, upload_file_id, idx, merchant_id, "fos_not_found", f"FOS '{gse}' not found in system")
                    error_count += 1
                    continue

                email_id = str(_get(row_vals, "Email ID") or "").strip() or None
                if email_id == "None":
                    email_id = None
                remark = str(_get(row_vals, "Type") or "").strip() or None
                if remark == "None":
                    remark = None

                assigned_on_raw = _get(row_vals, "Assigned on")
                doa = None
                if assigned_on_raw:
                    try:
                        if isinstance(assigned_on_raw, (date, datetime)):
                            doa = assigned_on_raw if isinstance(assigned_on_raw, date) else assigned_on_raw.date()
                        else:
                            from datetime import datetime as dt
                            doa = dt.strptime(str(assigned_on_raw).strip(), "%d/%m/%Y").date()
                    except Exception:
                        doa = None

                entry_fields = {
                    "rnc_entry_date": None, "idv_entry_date": None, "rtl_entry_date": None,
                    "fba_entry_date": None, "sp_entry_date": None, "open_spending_entry_date": None,
                    "narf_entry_date": None, "gsi_entry_date": None,
                }
                for key, field in {
                    "rnc": "rnc_entry_date", "idv": "idv_entry_date", "rtl": "rtl_entry_date",
                    "fba": "fba_entry_date", "sp": "sp_entry_date", "open spending": "open_spending_entry_date",
                    "narf": "narf_entry_date", "gsi": "gsi_entry_date",
                }.items():
                    if key in (activity_name or "").lower():
                        entry_fields[field] = date.today()
                        break

                custom_data = {h: _json_safe(_get(row_vals, h)) for h in _extra_headers}

                leads_to_insert.append({
                    "id": str(uuid.uuid4()),
                    "merchant_id": merchant_id,
                    "seller_name": None,
                    "mobile_number": mobile,
                    "email_id": email_id,
                    "stage_assigned": activity_name,
                    "date_of_assignment": doa,
                    "week_no": None,
                    "year": doa.year if doa else None,
                    "current_activity_id": activity_id,
                    "current_stage": activity_name,
                    "assigned_fos_id": fos_id,
                    "remark": remark,
                    "source_upload_id": upload_file_id,
                    "custom_data": json.dumps(custom_data),
                    **entry_fields,
                })

            if leads_to_insert:
                session.execute(
                    text("""
                        INSERT INTO leads (
                            id, merchant_id, seller_name, mobile_number, email_id,
                            stage_assigned, date_of_assignment, week_no, year,
                            current_activity_id, current_stage, assigned_fos_id, remark,
                            source_upload_id, custom_data,
                            rnc_entry_date, idv_entry_date, rtl_entry_date, fba_entry_date,
                            sp_entry_date, open_spending_entry_date, narf_entry_date, gsi_entry_date
                        ) VALUES (
                            :id, :merchant_id, :seller_name, :mobile_number, :email_id,
                            :stage_assigned, :date_of_assignment, :week_no, :year,
                            :current_activity_id, :current_stage, :assigned_fos_id, :remark,
                            :source_upload_id, CAST(:custom_data AS JSONB),
                            :rnc_entry_date, :idv_entry_date, :rtl_entry_date, :fba_entry_date,
                            :sp_entry_date, :open_spending_entry_date, :narf_entry_date, :gsi_entry_date
                        )
                    """),
                    leads_to_insert,
                )
                success_count = len(leads_to_insert)

            _update_upload_status(session, upload_file_id, "completed", total, success_count, error_count)
            session.commit()

            if leads_to_insert:
                unique_fos_ids = list({lead["assigned_fos_id"] for lead in leads_to_insert})
                _publish_upload_complete(upload_file_id, success_count, error_count, unique_fos_ids)

            return {"total": total, "success": success_count, "errors": error_count}

        except Exception as exc:
            session.rollback()
            session.execute(
                text("UPDATE upload_files SET status = 'failed' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()
            raise exc


@celery_app.task(bind=True, name="process_call_upload")
def process_call_upload(self, upload_file_id: str) -> dict:
    """Parse call log file (Sheet1): aggregates by MCID → call_count and total_call_time (mins)."""
    with SyncSession() as session:
        try:
            row = session.execute(
                text("SELECT s3_key, admin_id FROM upload_files WHERE id = :id"),
                {"id": upload_file_id},
            ).fetchone()
            if not row:
                return {"error": "Upload record not found"}

            session.execute(
                text("UPDATE upload_files SET status = 'processing' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()

            file_bytes = _get_file_bytes(upload_file_id, row.s3_key)

            import openpyxl, io
            from collections import defaultdict
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            # Find sheet with MCID and Call time (In min) headers
            ws = None
            for sheet_name in wb.sheetnames:
                s = wb[sheet_name]
                headers = [str(c.value).strip() if c.value else "" for c in next(s.iter_rows(max_row=1))]
                if "MCID" in headers and "Call time (In min)" in headers:
                    ws = s
                    break
            if ws is None:
                session.execute(text("UPDATE upload_files SET status = 'failed' WHERE id = :id"), {"id": upload_file_id})
                session.commit()
                return {"error": "No sheet with MCID and Call time (In min) columns found"}

            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            mcid_idx = headers.index("MCID")
            calltime_idx = headers.index("Call time (In min)")

            # Aggregate per MCID
            agg = defaultdict(lambda: {"count": 0, "time": 0.0})
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_raw = len(all_rows)

            for row_vals in all_rows:
                mid = row_vals[mcid_idx]
                ct = row_vals[calltime_idx]
                if not mid or mid == 0:
                    continue
                merchant_id = str(int(mid)) if isinstance(mid, float) else str(mid).strip()
                try:
                    call_time = float(ct) if ct not in (None, "", "#VALUE!") else 0.0
                except (ValueError, TypeError):
                    call_time = 0.0
                agg[merchant_id]["count"] += 1
                agg[merchant_id]["time"] = round(agg[merchant_id]["time"] + call_time, 2)

            total = len(agg)
            success_count = 0
            error_count = 0
            errors_to_insert = []

            for merchant_id, vals in agg.items():
                lead_row = session.execute(
                    text("SELECT id FROM leads WHERE merchant_id = :mid"),
                    {"mid": merchant_id},
                ).fetchone()

                if not lead_row:
                    errors_to_insert.append({
                        "id": str(uuid.uuid4()),
                        "upload_file_id": upload_file_id,
                        "row_number": 0,
                        "merchant_id": merchant_id,
                        "error_type": "lead_not_found",
                        "error_detail": f"MCID {merchant_id} not found in system",
                    })
                    error_count += 1
                    continue

                session.execute(
                    text("""
                        UPDATE leads
                        SET call_count = call_count + :count,
                            total_call_time = total_call_time + :time
                        WHERE merchant_id = :mid
                    """),
                    {"count": vals["count"], "time": vals["time"], "mid": merchant_id},
                )
                success_count += 1

            if errors_to_insert:
                session.execute(
                    text("""
                        INSERT INTO upload_errors (id, upload_file_id, row_number, merchant_id, error_type, error_detail)
                        VALUES (:id, :upload_file_id, :row_number, :merchant_id, :error_type, :error_detail)
                    """),
                    errors_to_insert,
                )

            _update_upload_status(session, upload_file_id, "completed", total, success_count, error_count)
            session.commit()
            return {"total": total, "success": success_count, "errors": error_count}

        except Exception as exc:
            session.rollback()
            session.execute(
                text("UPDATE upload_files SET status = 'failed' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()
            raise exc


@celery_app.task(bind=True, name="process_template2_upload")
def process_template2_upload(self, upload_file_id: str) -> dict:
    with SyncSession() as session:
        try:
            row = session.execute(
                text("SELECT s3_key, admin_id FROM upload_files WHERE id = :id"),
                {"id": upload_file_id},
            ).fetchone()
            if not row:
                return {"error": "Upload record not found"}

            s3_key, admin_id = row.s3_key, row.admin_id

            session.execute(
                text("UPDATE upload_files SET status = 'processing' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()

            file_bytes = _get_file_bytes(upload_file_id, s3_key)
            _, rows = read_excel_rows(file_bytes)

            error_count = 0
            success_count = 0
            total = len(rows)

            for idx, row_data in enumerate(rows, start=2):
                merchant_id = str(row_data.get("Merchant ID") or "").strip()
                final_stage = str(row_data.get("Final Stage") or "").strip()
                week_no_raw = row_data.get("Week No")
                year_raw = row_data.get("Year")
                remark = str(row_data.get("Remark") or "").strip() or None

                if not merchant_id:
                    _insert_error(session, upload_file_id, idx, None, "missing_field", "Merchant ID is blank")
                    error_count += 1
                    continue

                if not final_stage:
                    _insert_error(session, upload_file_id, idx, merchant_id, "missing_field", "Final Stage is blank")
                    error_count += 1
                    continue

                lead_row = session.execute(
                    text("SELECT id, current_activity_id FROM leads WHERE merchant_id = :mid"),
                    {"mid": merchant_id},
                ).fetchone()

                if not lead_row:
                    _insert_error(session, upload_file_id, idx, merchant_id, "lead_not_found", f"Merchant ID {merchant_id} not found in system")
                    error_count += 1
                    continue

                lead_id = lead_row.id
                current_activity_id = lead_row.current_activity_id
                week_no = int(week_no_raw) if week_no_raw else None
                year = int(year_raw) if year_raw else None

                # Find next activity
                if current_activity_id:
                    order_result = session.execute(
                        text("SELECT position_order FROM activities WHERE id = :id"),
                        {"id": str(current_activity_id)},
                    ).fetchone()
                    current_order = order_result[0] if order_result else None
                    next_act = None
                    if current_order is not None:
                        next_act_result = session.execute(
                            text("SELECT id, name FROM activities WHERE position_order > :order AND is_active = true ORDER BY position_order LIMIT 1"),
                            {"order": current_order},
                        ).fetchone()
                        next_act = next_act_result

                    entry_field = None
                    if next_act:
                        name_lower = next_act.name.lower()
                        for key, field in {
                            "rnc": "rnc_entry_date", "idv": "idv_entry_date",
                            "rtl": "rtl_entry_date", "fba": "fba_entry_date",
                            "sp": "sp_entry_date", "open spending": "open_spending_entry_date",
                            "narf": "narf_entry_date", "gsi": "gsi_entry_date",
                        }.items():
                            if key in name_lower:
                                entry_field = field
                                break

                    update_fields = {
                        "final_stage": final_stage,
                        "week_of_movement": week_no,
                        "sub_disposition": None,
                        "follow_up_date": None,
                        "updated_at": datetime.now(timezone.utc),
                    }
                    if next_act:
                        update_fields["current_activity_id"] = str(next_act.id)
                        update_fields["current_stage"] = next_act.name
                        if entry_field:
                            update_fields[entry_field] = date.today()

                    if remark:
                        update_fields["remark"] = remark

                    set_clause = ", ".join(f"{k} = :{k}" for k in update_fields)
                    update_fields["lead_id"] = str(lead_id)
                    session.execute(
                        text(f"UPDATE leads SET {set_clause} WHERE id = :lead_id"),
                        update_fields,
                    )

                    # Write history
                    session.execute(
                        text("""
                            INSERT INTO lead_history (id, lead_id, action_type, old_value, new_value, performed_by)
                            VALUES (:id, :lead_id, :action_type, :old_value, :new_value, :performed_by)
                        """),
                        {
                            "id": str(uuid.uuid4()),
                            "lead_id": str(lead_id),
                            "action_type": "movement",
                            "old_value": str(current_activity_id),
                            "new_value": str(next_act.id) if next_act else None,
                            "performed_by": str(admin_id),
                        },
                    )
                    success_count += 1

            _update_upload_status(session, upload_file_id, "completed", total, success_count, error_count)
            session.commit()
            return {"total": total, "success": success_count, "errors": error_count}

        except Exception as exc:
            session.rollback()
            session.execute(
                text("UPDATE upload_files SET status = 'failed' WHERE id = :id"),
                {"id": upload_file_id},
            )
            session.commit()
            raise exc
