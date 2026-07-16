import uuid
import logging
from datetime import datetime, timezone

from openpyxl import load_workbook
import io

from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

from app.celery_app import celery_app
from app.config import settings
from app.services.s3_service import download_file_from_s3
from app.services.upload_service import get_dev_temp_path

logger = logging.getLogger(__name__)

_sync_db_url = settings.database_url.replace("postgresql+asyncpg", "postgresql+psycopg2")
_sync_engine = create_engine(_sync_db_url, pool_pre_ping=True)
SyncSession = sessionmaker(bind=_sync_engine)


def _get_file_bytes(upload_id: str, s3_key: str) -> bytes:
    if not settings.r2_bucket:
        path = get_dev_temp_path(upload_id)
        with open(path, "rb") as f:
            return f.read()
    return download_file_from_s3(s3_key)


def _finish(session: Session, upload_id: str, status: str, total: int, success: int, errors: int) -> None:
    session.execute(
        text("""
            UPDATE upload_files
            SET status=:s, total_rows=:t, success_rows=:ok, error_rows=:err, completed_at=:now
            WHERE id=:id
        """),
        {"s": status, "t": total, "ok": success, "err": errors,
         "now": datetime.now(timezone.utc), "id": upload_id},
    )


@celery_app.task(bind=True, name="process_campaign_upload")
def process_campaign_upload(self, upload_file_id: str, campaign_id: str) -> dict:
    with SyncSession() as session:
        try:
            row_result = session.execute(
                text("SELECT s3_key FROM upload_files WHERE id=:id"),
                {"id": upload_file_id},
            ).fetchone()
            if not row_result:
                logger.error("upload_file not found: %s", upload_file_id)
                return {"status": "error"}

            file_bytes = _get_file_bytes(upload_file_id, row_result[0])
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active

            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            try:
                mcid_col = next(i for i, h in enumerate(headers) if "mcid" in h.lower() or "merchant" in h.lower())
            except StopIteration:
                _finish(session, upload_file_id, "failed", 0, 0, 0)
                session.commit()
                return {"status": "error", "detail": "No MCID column found"}

            remark_col = next(
                (i for i, h in enumerate(headers) if "remark" in h.lower() or "note" in h.lower()),
                None
            )

            success, errors = 0, 0
            rows_data = list(ws.iter_rows(min_row=2, values_only=True))
            total = len([r for r in rows_data if any(c is not None for c in r)])

            for row in rows_data:
                if all(c is None for c in row):
                    continue
                raw_mid = row[mcid_col]
                if raw_mid is None:
                    errors += 1
                    continue

                mid = str(int(raw_mid)) if isinstance(raw_mid, float) else str(raw_mid).strip()
                remark = str(row[remark_col]).strip() if remark_col is not None and row[remark_col] is not None else None

                lead_row = session.execute(
                    text("SELECT id FROM leads WHERE merchant_id=:mid"),
                    {"mid": mid},
                ).fetchone()

                if not lead_row:
                    errors += 1
                    continue

                # Skip duplicates within same campaign
                existing = session.execute(
                    text("SELECT id FROM campaign_leads WHERE campaign_id=:cid AND lead_id=:lid"),
                    {"cid": campaign_id, "lid": str(lead_row[0])},
                ).fetchone()
                if existing:
                    continue

                session.execute(
                    text("""
                        INSERT INTO campaign_leads (id, campaign_id, lead_id, merchant_id, event_remark)
                        VALUES (:id, :cid, :lid, :mid, :remark)
                    """),
                    {
                        "id": str(uuid.uuid4()),
                        "cid": campaign_id,
                        "lid": str(lead_row[0]),
                        "mid": mid,
                        "remark": remark,
                    },
                )
                success += 1

            _finish(session, upload_file_id, "complete", total, success, errors)
            session.commit()
            logger.info("Campaign upload %s done: %d ok, %d err", upload_file_id, success, errors)
            return {"status": "complete", "success": success, "errors": errors}

        except Exception as exc:
            session.rollback()
            logger.exception("Campaign upload failed: %s", exc)
            _finish(session, upload_file_id, "failed", 0, 0, 0)
            session.commit()
            return {"status": "error"}
