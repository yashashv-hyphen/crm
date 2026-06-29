import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.dependencies import get_current_user
from app.models.user import User
from app.utils.excel_validator import TEMPLATE1_REQUIRED_COLUMNS, TEMPLATE2_REQUIRED_COLUMNS

router = APIRouter(prefix="/api/templates", tags=["templates"])


def _make_excel(headers: list[str], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/template1")
async def download_template1(_: User = Depends(get_current_user)):
    extra = ["Email ID", "Remark"]
    all_headers = TEMPLATE1_REQUIRED_COLUMNS + extra
    excel_bytes = _make_excel(all_headers, "New Leads Upload")
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template1_new_leads.xlsx"},
    )


@router.get("/template2")
async def download_template2(_: User = Depends(get_current_user)):
    extra = ["Remark"]
    all_headers = TEMPLATE2_REQUIRED_COLUMNS + extra
    excel_bytes = _make_excel(all_headers, "Final Stage Update")
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template2_final_stage.xlsx"},
    )
