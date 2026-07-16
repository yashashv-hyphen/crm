import io
from openpyxl import load_workbook

TEMPLATE1_REQUIRED_COLUMNS = [
    "Stage Assigned", "Date of Assignment", "Week No", "Year",
    "Merchant ID", "Mobile Number", "Seller Name", "Assigned To FOS",
]

TEMPLATE2_REQUIRED_COLUMNS = [
    "Merchant ID", "Final Stage", "Week No", "Year",
]

REGULAR_REQUIRED_COLUMNS = [
    "merchant_customer_id", "Lead Stage", "GSE",
]

MAX_ROWS = 5000


def validate_template1_pre_upload(file_bytes: bytes) -> list[str]:
    """Stage 1 validation — returns list of error strings; empty = OK."""
    errors = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return ["File is not a valid .xlsx file"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in TEMPLATE1_REQUIRED_COLUMNS:
        if col not in headers:
            errors.append(f"Missing required column: '{col}'")

    row_count = ws.max_row - 1  # subtract header row
    if row_count > MAX_ROWS:
        errors.append(f"File has {row_count} rows. Maximum allowed is {MAX_ROWS}.")

    return errors


def validate_template2_pre_upload(file_bytes: bytes) -> list[str]:
    errors = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return ["File is not a valid .xlsx file"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in TEMPLATE2_REQUIRED_COLUMNS:
        if col not in headers:
            errors.append(f"Missing required column: '{col}'")

    row_count = ws.max_row - 1
    if row_count > MAX_ROWS:
        errors.append(f"File has {row_count} rows. Maximum allowed is {MAX_ROWS}.")

    return errors


def validate_call_pre_upload(file_bytes: bytes) -> list[str]:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return ["File is not a valid .xlsx file"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        if "MCID" in headers and "Call time (In min)" in headers:
            return []
    return ["No sheet found with both 'MCID' and 'Call time (In min)' columns"]


def validate_regular_pre_upload(file_bytes: bytes) -> list[str]:
    errors = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return ["File is not a valid .xlsx file"]

    # Find any sheet containing merchant_customer_id header
    found_sheet = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        if "merchant_customer_id" in headers:
            found_sheet = ws
            break

    if found_sheet is None:
        return ["No sheet found with 'merchant_customer_id' column"]

    headers = [cell.value for cell in next(found_sheet.iter_rows(max_row=1))]
    for col in REGULAR_REQUIRED_COLUMNS:
        if col not in headers:
            errors.append(f"Missing required column: '{col}'")

    row_count = found_sheet.max_row - 1
    if row_count > MAX_ROWS:
        errors.append(f"File has {row_count} rows. Maximum allowed is {MAX_ROWS}.")

    return errors


def read_excel_rows(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    """Returns (headers, rows_as_dicts)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        data.append(dict(zip(headers, row)))
    return headers, data
